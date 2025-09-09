from datetime import datetime
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file
from flask_login import login_required, current_user
from app.models import Department, MetricExclusion, db
from app.models import (
    ContactMessage, Employee, EvaluationCycle, MetricCategory,
    PerformanceMetric, EmployeeMetric, Feedback, FeedbackType,
    FAQ, News, Role, Department, Position
)
from fpdf import FPDF
from io import BytesIO
from fpdf.enums import XPos, YPos

views = Blueprint('views', __name__)
FONT_PATH = "app/static/fonts/DejaVuSans.ttf"
FONT_PATH_B = "app/static/fonts/DejaVuSans-Bold.ttf"
FONT_PATH_I = "app/static/fonts/DejaVuSans-BoldOblique.ttf"

# Вспомогательная функция для хлебных крошек
def render_with_breadcrumbs(template, breadcrumbs, **context):
    return render_template(template, breadcrumbs=breadcrumbs, **context)

# Проверка прав: только admin и manager
def admin_or_manager_required(f):
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            abort(403)
        if current_user.role.name not in ['admin', 'manager']:
            abort(403)
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@views.route('/')
@views.route('/index')
def index():
    from datetime import datetime
    current_date = datetime.utcnow()
    expired_active_cycles = EvaluationCycle.query.filter(
        EvaluationCycle.is_active == True,
        EvaluationCycle.end_date < current_date
    ).all()
    for cycle in expired_active_cycles:
        cycle.is_active = False
        db.session.add(cycle)
    if expired_active_cycles:
        db.session.commit()
    active_cycles = EvaluationCycle.query.filter_by(is_active=True).all()
    latest_news = News.query.filter_by(is_published=True).order_by(News.published_at.desc()).limit(3).all()
    breadcrumbs = [("Главная", url_for('views.index'))]
    return render_with_breadcrumbs('index.html', breadcrumbs, active_cycles=active_cycles, latest_news=latest_news)

@views.route('/about')
def about():
    breadcrumbs = [("Главная", url_for('views.index')), ("О компании", url_for('views.about'))]
    return render_with_breadcrumbs('about.html', breadcrumbs)

@views.route('/contact', methods=['GET', 'POST'])
def contact():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        message = request.form.get('message')
        if not name or not email or not message:
            flash('Заполните все поля.', 'error')
        else:
            msg = ContactMessage(name=name, email=email, message=message)
            db.session.add(msg)
            db.session.commit()
            flash('Сообщение отправлено! Спасибо.', 'success')
            return redirect(url_for('views.contact'))
    breadcrumbs = [("Главная", url_for('views.index')), ("Контакты", url_for('views.contact'))]
    return render_with_breadcrumbs('contact.html', breadcrumbs)

@views.route('/admin/employees')
@login_required
@admin_or_manager_required
def admin_employees():
    dept_filter = request.args.get('department', type=int)
    if current_user.role.name == 'manager':
        query = Employee.query.filter_by(department_id=current_user.department_id)
    else:
        query = Employee.query
    if dept_filter:
        query = query.filter_by(department_id=dept_filter)
    employees_list = query.all()
    if current_user.role.name == 'manager':
        departments = [current_user.department]
    else:
        departments = Department.query.all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Управление сотрудниками", "")]
    return render_with_breadcrumbs('admin_employees.html', breadcrumbs, employees=employees_list, departments=departments, selected_dept=dept_filter)

@views.route('/employees')
def employees():
    dept_filter = request.args.get('department', type=int)
    query = Employee.query.filter_by(is_active=True)
    if dept_filter:
        query = query.filter_by(department_id=dept_filter)
    emps = query.all()
    departments = Department.query.all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Сотрудники", url_for('views.employees'))]
    return render_with_breadcrumbs('employees.html', breadcrumbs, employees=emps, departments=departments, selected_dept=dept_filter)

@views.route('/metrics')
def metrics():
    categories = MetricCategory.query.all()
    metrics = PerformanceMetric.query.filter_by(is_active=True).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Метрики оценки", url_for('views.metrics'))]
    return render_with_breadcrumbs('metrics.html', breadcrumbs, categories=categories, metrics=metrics)

@views.route('/cycles')
def cycles():
    from datetime import datetime
    current_date = datetime.utcnow()
    expired_active_cycles = EvaluationCycle.query.filter(
        EvaluationCycle.is_active == True,
        EvaluationCycle.end_date < current_date
    ).all()
    for cycle in expired_active_cycles:
        cycle.is_active = False
        db.session.add(cycle)
    if expired_active_cycles:
        db.session.commit()
    cycles_list = EvaluationCycle.query.order_by(EvaluationCycle.start_date.desc()).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Циклы оценки", url_for('views.cycles'))]
    return render_with_breadcrumbs('cycles.html', breadcrumbs, cycles=cycles_list)

@views.route('/faq')
def faq():
    faq_items = FAQ.query.filter_by(is_published=True).order_by(FAQ.category, FAQ.id).all()
    categories = sorted(set(item.category for item in faq_items))
    breadcrumbs = [("Главная", url_for('views.index')), ("Частые вопросы", url_for('views.faq'))]
    return render_with_breadcrumbs('faq.html', breadcrumbs, faq_items=faq_items, categories=categories)

@views.route('/documentation')
def documentation():
    breadcrumbs = [("Главная", url_for('views.index')), ("Документация", url_for('views.documentation'))]
    return render_with_breadcrumbs('documentation.html', breadcrumbs)

@views.route('/news')
def news_list():
    news_items = News.query.filter_by(is_published=True).order_by(News.published_at.desc()).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Новости", url_for('views.news_list'))]
    return render_with_breadcrumbs('news.html', breadcrumbs, news_items=news_items)

@views.route('/news/<int:news_id>')
def news_detail(news_id):
    item = News.query.filter_by(id=news_id, is_published=True).first_or_404()
    breadcrumbs = [
        ("Главная", url_for('views.index')),
        ("Новости", url_for('views.news_list')),
        (item.title, url_for('views.news_detail', news_id=item.id))
    ]
    return render_with_breadcrumbs('news_detail.html', breadcrumbs, news=item)

@views.route('/stats')
def stats():
    total_emps = Employee.query.filter_by(is_active=True).count()
    total_cycles = EvaluationCycle.query.count()
    active_metrics = PerformanceMetric.query.filter_by(is_active=True).count()
    feedback_count = Feedback.query.count()

    active_cycle = EvaluationCycle.query.filter_by(is_active=True).first()

    if not active_cycle:
        flash("Нет активного цикла оценки.", "info")
        top_employees = []
        evaluated_count = 0
        not_evaluated_count = 0
        department_stats = []
        category_names = []
        category_scores = []
        all_scores = []
        dept_names = []
        dept_avg_scores = []
    else:
        from sqlalchemy import func

        # --- Топ-5 ---
        top_employee_data = db.session.query(
            Employee.full_name,
            Department.name.label('dept_name'),
            func.avg(EmployeeMetric.score).label('avg_score')
        ).join(EmployeeMetric, Employee.id == EmployeeMetric.employee_id) \
         .join(Department, Employee.department_id == Department.id) \
         .filter(EmployeeMetric.cycle_id == active_cycle.id) \
         .group_by(Employee.id, Department.name) \
         .order_by(func.avg(EmployeeMetric.score).desc()) \
         .limit(5).all()

        top_employees = [
            {"full_name": row.full_name, "department": {"name": row.dept_name}, "avg_score": float(row.avg_score)}
            for row in top_employee_data
        ]

        # --- Завершённость ---
        all_employees = Employee.query.filter_by(is_active=True).all()
        evaluated_ids = {item[0] for item in db.session.query(EmployeeMetric.employee_id).filter(
            EmployeeMetric.cycle_id == active_cycle.id
        ).distinct().all()}
        evaluated_count = len(evaluated_ids)
        not_evaluated_count = len(all_employees) - evaluated_count

        # --- Активность по подразделениям ---
        department_stats = []
        for dept in Department.query.all():
            emps = Employee.query.filter_by(department_id=dept.id, is_active=True).all()
            if emps:
                total = len(emps)
                evaluated = len([e for e in emps if e.id in evaluated_ids])
                percent = (evaluated / total) * 100
                department_stats.append({
                    "name": dept.name,
                    "total": total,
                    "evaluated": evaluated,
                    "percent": round(percent, 1)
                })

        # --- Все оценки для гистограммы ---
        all_scores = [float(row.score) for row in db.session.query(EmployeeMetric.score).filter(
            EmployeeMetric.cycle_id == active_cycle.id
        ).all()]

        # --- Средний балл по подразделениям (для pie и bar) ---
        dept_avg_data = db.session.query(
            Department.name,
            func.avg(EmployeeMetric.score).label('avg_score')
        ).join(Employee, Department.id == Employee.department_id) \
         .join(EmployeeMetric, Employee.id == EmployeeMetric.employee_id) \
         .filter(EmployeeMetric.cycle_id == active_cycle.id) \
         .group_by(Department.id, Department.name).all()

        dept_names = [row.name for row in dept_avg_data]
        dept_avg_scores = [float(row.avg_score) for row in dept_avg_data]

        # --- Средние по категориям ---
        category_data = db.session.query(
            MetricCategory.name,
            func.avg(EmployeeMetric.score).label('avg_score')
        ).join(PerformanceMetric, EmployeeMetric.metric_id == PerformanceMetric.id) \
        .join(MetricCategory, PerformanceMetric.category_id == MetricCategory.id) \
        .filter(EmployeeMetric.cycle_id == active_cycle.id) \
        .group_by(MetricCategory.id, MetricCategory.name) \
        .all()

        category_names = [item.name for item in category_data]
        category_scores = [float(item.avg_score) if item.avg_score else 0.0 for item in category_data]

    breadcrumbs = [("Главная", url_for('views.index')), ("Статистика", url_for('views.stats'))]
    return render_with_breadcrumbs('stats.html', breadcrumbs,
                                   total_emps=total_emps,
                                   total_cycles=total_cycles,
                                   active_metrics=active_metrics,
                                   feedback_count=feedback_count,
                                   active_cycle=active_cycle,
                                   top_employees=top_employees,
                                   evaluated_count=evaluated_count,
                                   not_evaluated_count=not_evaluated_count,
                                   department_stats=department_stats,
                                   category_names=category_names,
                                   category_scores=category_scores,
                                   all_scores=all_scores,
                                   dept_names=dept_names,
                                   dept_avg_scores=dept_avg_scores)

@views.route('/stats/all-employees')
def all_employees():
    active_cycle = EvaluationCycle.query.filter_by(is_active=True).first()
    if not active_cycle:
        flash("Нет активного цикла оценки.", "warning")
        return redirect(url_for('views.stats'))

    from sqlalchemy import func

    # Все сотрудники с их средними баллами
    all_employee_data = db.session.query(
        Employee.full_name,
        Department.name.label('dept_name'),
        func.avg(EmployeeMetric.score).label('avg_score')
    ).join(EmployeeMetric, Employee.id == EmployeeMetric.employee_id) \
     .join(Department, Employee.department_id == Department.id) \
     .filter(EmployeeMetric.cycle_id == active_cycle.id) \
     .group_by(Employee.id, Department.name) \
     .order_by(func.avg(EmployeeMetric.score).desc()) \
     .all()

    employees = [
        {"full_name": row.full_name, "department": {"name": row.dept_name}, "avg_score": float(row.avg_score)}
        for row in all_employee_data
    ]

    breadcrumbs = [
        ("Главная", url_for('views.index')),
        ("Статистика", url_for('views.stats')),
        ("Все сотрудники", url_for('views.all_employees'))
    ]
    return render_with_breadcrumbs('all_employees.html', breadcrumbs, employees=employees, cycle_name=active_cycle.name)

@views.route('/categories')
def categories():
    cats = MetricCategory.query.all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Категории KPI", url_for('views.categories'))]
    return render_with_breadcrumbs('categories.html', breadcrumbs, categories=cats)

@views.route('/admin/faq')
@login_required
@admin_or_manager_required
def admin_faq():
    faqs = FAQ.query.order_by(FAQ.category, FAQ.id).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Админ: FAQ", url_for('views.admin_faq'))]
    return render_with_breadcrumbs('admin_faq.html', breadcrumbs, faqs=faqs)

@views.route('/admin/faq/add', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_faq():
    if request.method == 'POST':
        question = request.form.get('question')
        answer = request.form.get('answer')
        category = request.form.get('category', 'Общее')
        if question and answer:
            faq = FAQ(
                question=question,
                answer=answer,
                category=category,
                author_id=current_user.id
            )
            db.session.add(faq)
            db.session.commit()
            flash("Вопрос добавлен.", "success")
            return redirect(url_for('views.admin_faq'))
        else:
            flash("Заполните вопрос и ответ.", "error")
    breadcrumbs = [("Главная", url_for('views.index')), ("Добавить FAQ", "")]
    return render_with_breadcrumbs('add_faq.html', breadcrumbs)

@views.route('/admin/faq/edit/<int:faq_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_faq(faq_id):
    faq = FAQ.query.get_or_404(faq_id)
    if request.method == 'POST':
        faq.question = request.form.get('question')
        faq.answer = request.form.get('answer')
        faq.category = request.form.get('category', 'Общее')
        faq.updated_at = datetime.utcnow()
        db.session.commit()
        flash("Вопрос обновлён.", "success")
        return redirect(url_for('views.admin_faq'))
    breadcrumbs = [("Главная", url_for('views.index')), ("Редактировать FAQ", "")]
    return render_with_breadcrumbs('edit_faq.html', breadcrumbs, faq=faq)

@views.route('/admin/faq/delete/<int:faq_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_faq(faq_id):
    faq = FAQ.query.get_or_404(faq_id)
    db.session.delete(faq)
    db.session.commit()
    flash("Вопрос удалён.", "success")
    return redirect(url_for('views.admin_faq'))

@views.route('/admin/news')
@login_required
@admin_or_manager_required
def admin_news():
    news_items = News.query.order_by(News.published_at.desc()).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Админ: Новости", url_for('views.admin_news'))]
    return render_with_breadcrumbs('admin_news.html', breadcrumbs, news_items=news_items)

@views.route('/admin/news/add', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_news():
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        published_at = request.form.get('published_at')
        image_url = request.form.get('image_url')

        if not title or not content:
            flash("Заполните заголовок и текст.", "error")
        else:
            news = News(
                title=title,
                content=content,
                published_at=datetime.fromisoformat(published_at) if published_at else datetime.utcnow(),
                image_url=image_url,
                author_id=current_user.id
            )
            db.session.add(news)
            db.session.commit()
            flash("Новость добавлена.", "success")
            return redirect(url_for('views.admin_news'))
    breadcrumbs = [("Главная", url_for('views.index')), ("Добавить новость", "")]
    return render_with_breadcrumbs('add_news.html', breadcrumbs)

@views.route('/admin/news/edit/<int:news_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_news(news_id):
    news = News.query.get_or_404(news_id)
    if request.method == 'POST':
        news.title = request.form.get('title')
        news.content = request.form.get('content')
        news.published_at = datetime.fromisoformat(request.form.get('published_at'))
        news.image_url = request.form.get('image_url')
        news.updated_at = datetime.utcnow()
        db.session.commit()
        flash("Новость обновлена.", "success")
        return redirect(url_for('views.admin_news'))
    breadcrumbs = [("Главная", url_for('views.index')), ("Редактировать новость", "")]
    return render_with_breadcrumbs('edit_news.html', breadcrumbs, news=news)

@views.route('/admin/news/delete/<int:news_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_news(news_id):
    news = News.query.get_or_404(news_id)
    db.session.delete(news)
    db.session.commit()
    flash("Новость удалена.", "success")
    return redirect(url_for('views.admin_news'))

# --- Защищённые страницы (для всех авторизованных) ---

@views.route('/dashboard')
@login_required
def dashboard():
    user = current_user
    received_evals = EmployeeMetric.query.filter_by(employee_id=user.id)\
        .join(EvaluationCycle).filter(EvaluationCycle.is_active == True).all()
    given_evals = EmployeeMetric.query.filter_by(evaluator_id=user.id)\
        .join(EvaluationCycle).filter(EvaluationCycle.is_active == True).all()
    feedbacks = Feedback.query.filter_by(employee_id=user.id).all()

    breadcrumbs = [("Главная", url_for('views.index')), ("Личный кабинет", url_for('views.dashboard'))]
    return render_with_breadcrumbs('dashboard.html', breadcrumbs,
                                   user=user,
                                   received_evals=received_evals,
                                   given_evals=given_evals,
                                   feedbacks=feedbacks)

@views.route('/feedback/send', methods=['GET', 'POST'])
@login_required
def send_feedback():
    employees = Employee.query.filter(Employee.id != current_user.id).all()
    types = FeedbackType.query.all()

    if request.method == 'POST':
        emp_id = request.form.get('employee_id')
        type_id = request.form.get('type_id')
        content = request.form.get('content')
        is_anon = bool(request.form.get('anonymous'))

        if not emp_id or not type_id or not content:
            flash("Заполните все поля.", "error")
        else:
            fb = Feedback(
                employee_id=int(emp_id),
                sender_id=current_user.id,
                feedback_type_id=int(type_id),
                content=content,
                is_anonymous=is_anon
            )
            db.session.add(fb)
            db.session.commit()
            flash("Обратная связь отправлена.", "success")
            return redirect(url_for('views.dashboard'))

    breadcrumbs = [("Главная", url_for('views.index')), ("Отправить обратную связь", "")]
    return render_with_breadcrumbs('send_feedback.html', breadcrumbs,
                                   employees=employees, types=types)

@views.route('/admin/metrics')
@login_required
@admin_or_manager_required
def admin_metrics():
    if current_user.role.name == 'manager':
        metrics = PerformanceMetric.query.filter_by(department_id=current_user.department_id).all()
    else:
        metrics = PerformanceMetric.query.all()

    categories = MetricCategory.query.all()
    departments = Department.query.all()

    breadcrumbs = [("Главная", url_for('views.index')), ("Управление метриками", "")]
    return render_with_breadcrumbs('admin_metrics.html', breadcrumbs,
                                   metrics=metrics,
                                   categories=categories,
                                   departments=departments)

@views.route('/admin/cycles')
@login_required
@admin_or_manager_required
def admin_cycles():
    cycles = EvaluationCycle.query.order_by(EvaluationCycle.start_date.desc()).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Управление циклами", "")]
    return render_with_breadcrumbs('admin_cycles.html', breadcrumbs, cycles=cycles)

@views.route('/admin/metrics/add', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_metric():
    if request.method == 'POST':
        name = request.form.get('name')
        category_id = request.form.get('category_id')
        max_score = float(request.form.get('max_score', 10))
        weight = float(request.form.get('weight', 1.0))
        description = request.form.get('description')
        department_id = request.form.get('department_id')  # Может быть None

        # Менеджер может добавить только в своё подразделение
        if current_user.role.name == 'manager':
            if department_id and int(department_id) != current_user.department_id:
                flash("Вы можете добавлять метрики только в своё подразделение.", "error")
                return redirect(url_for('views.admin_metrics'))

        metric = PerformanceMetric(
            name=name,
            category_id=category_id,
            max_score=max_score,
            weight=weight,
            description=description,
            department_id=department_id if department_id else None
        )
        db.session.add(metric)
        db.session.flush()
        excluded_position_ids = request.form.getlist('exclude_position_id')
        for pos_id in excluded_position_ids:
            if pos_id:
                exclusion = MetricExclusion(metric_id=metric.id, position_id=int(pos_id))
                db.session.add(exclusion)
        excluded_employee_ids = request.form.getlist('exclude_employee_id')
        for emp_id in excluded_employee_ids:
            if emp_id:
                exclusion = MetricExclusion(metric_id=metric.id, employee_id=int(emp_id))
                db.session.add(exclusion)
        db.session.commit()
        flash("Метрика добавлена.", "success")
        return redirect(url_for('views.admin_metrics'))

    categories = MetricCategory.query.all()
    departments = [current_user.department] if current_user.role.name == 'manager' else Department.query.all()
    all_positions = Position.query.all()
    all_employees = Employee.query.filter_by(is_active=True).all()

    return render_template('add_metric.html',
                           categories=categories,
                           departments=departments,
                           all_positions=all_positions,
                           all_employees=all_employees)

# --- Управление сотрудниками ---
@views.route('/admin/employees/add', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_employee():
    departments = Department.query.all()
    roles = Role.query.all()
    positions = Position.query.all()

    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        department_id = request.form.get('department_id')
        position_id = request.form.get('position_id')
        role_id = request.form.get('role_id')
        ip_address = request.form.get('ip_address')
        password = request.form.get('password')

        if not all([full_name, email, department_id, position_id, role_id, password]):
            flash("Заполните все обязательные поля.", "error")
        elif Employee.query.filter_by(email=email).first():
            flash("Сотрудник с таким email уже существует.", "error")
        else:
            import re
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                flash("Некорректный формат email-адреса.", "error")
                return redirect(url_for('views.add_employee'))

            if ip_address:
                ip_regex = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
                if not re.match(ip_regex, ip_address):
                    flash("Некорректный формат IP-адреса. Используйте формат: 192.168.1.1", "error")
                    return redirect(url_for('views.add_employee'))

            target_role = Role.query.get(int(role_id))
            if target_role.name == 'admin' and current_user.role.name != 'admin':
                flash("Только администратор может создавать других администраторов.", "error")
                return redirect(url_for('views.add_employee'))

            from werkzeug.security import generate_password_hash
            employee = Employee(
                full_name=full_name,
                email=email,
                password_hash=generate_password_hash(password),
                department_id=int(department_id),
                position_id=int(position_id),
                role_id=int(role_id),
                ip_address=ip_address,
                is_active=True
            )
            db.session.add(employee)
            db.session.commit()
            flash(f"Сотрудник {full_name} добавлен.", "success")
            return redirect(url_for('views.admin_employees'))

    breadcrumbs = [("Главная", url_for('views.index')), ("Добавить сотрудника", "")]
    return render_with_breadcrumbs('add_employee.html', breadcrumbs,
                                   departments=departments,
                                   roles=roles,
                                   positions=positions)


@views.route('/admin/employees/edit/<int:emp_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_employee(emp_id):
    if emp_id == current_user.id:
        flash("Вы не можете редактировать свой собственный профиль через этот интерфейс.", "error")
        return redirect(url_for('views.admin_employees'))

    employee = Employee.query.get_or_404(emp_id)
    departments = Department.query.all()
    roles = Role.query.all()
    positions = Position.query.filter_by(department_id=employee.department_id).all()

    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        department_id = request.form.get('department_id')
        position_id = request.form.get('position_id')
        role_id = request.form.get('role_id')
        ip_address = request.form.get('ip_address')

        if not all([full_name, email, department_id, position_id, role_id]):
            flash("Заполните все обязательные поля.", "error")
        else:
            import re
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                flash("Некорректный формат email-адреса.", "error")
                return redirect(url_for('views.edit_employee', emp_id=emp_id))

            if ip_address:
                ip_regex = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
                if not re.match(ip_regex, ip_address):
                    flash("Некорректный формат IP-адреса. Используйте формат: 192.168.1.1", "error")
                    return redirect(url_for('views.edit_employee', emp_id=emp_id))

            target_role = Role.query.get(int(role_id))
            if target_role.name == 'admin' and current_user.role.name != 'admin':
                flash("Только администратор может назначать роль администратора.", "error")
                return redirect(url_for('views.edit_employee', emp_id=emp_id))

            employee.full_name = full_name
            employee.email = email
            employee.department_id = int(department_id)
            employee.position_id = int(position_id)
            employee.role_id = int(role_id)
            employee.ip_address = ip_address

            password = request.form.get('password')
            if password:
                from werkzeug.security import generate_password_hash
                employee.password_hash = generate_password_hash(password)

            db.session.commit()
            flash("Данные сотрудника обновлены.", "success")
            return redirect(url_for('views.admin_employees'))

    # Обновляем должности при смене подразделения (в шаблоне — JS)
    positions = Position.query.filter_by(department_id=employee.department_id).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Редактировать сотрудника", "")]
    return render_with_breadcrumbs('edit_employee.html', breadcrumbs,
                                   employee=employee,
                                   departments=departments,
                                   roles=roles,
                                   positions=positions)


@views.route('/admin/employees/delete/<int:emp_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_employee(emp_id):
    if emp_id == current_user.id:
        flash("Вы не можете удалить самого себя.", "error")
        return redirect(url_for('views.admin_employees'))

    employee = Employee.query.get_or_404(emp_id)
    if employee.role.name == 'admin' and current_user.role.name != 'admin':
        flash("Вы не можете удалить другого администратора.", "error")
    else:
        db.session.delete(employee)
        db.session.commit()
        flash("Сотрудник удалён.", "success")
    return redirect(url_for('views.admin_employees'))


@views.route('/admin/metrics/edit/<int:metric_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_metric(metric_id):
    metric = PerformanceMetric.query.get_or_404(metric_id)
    categories = MetricCategory.query.all()
    departments = Department.query.all()

    if (current_user.role.name == 'manager' and
        metric.department_id and
        metric.department_id != current_user.department_id):
        abort(403)

    if request.method == 'POST':
        metric.name = request.form.get('name')
        metric.description = request.form.get('description')
        metric.category_id = request.form.get('category_id')
        metric.max_score = float(request.form.get('max_score'))
        metric.weight = float(request.form.get('weight'))
        dept_id = request.form.get('department_id')
        metric.department_id = int(dept_id) if dept_id else None

        MetricExclusion.query.filter_by(metric_id=metric.id).delete()

        excluded_position_ids = request.form.getlist('exclude_position_id')
        for pos_id in excluded_position_ids:
            if pos_id:
                exclusion = MetricExclusion(metric_id=metric.id, position_id=int(pos_id))
                db.session.add(exclusion)
        excluded_employee_ids = request.form.getlist('exclude_employee_id')
        for emp_id in excluded_employee_ids:
            if emp_id:
                exclusion = MetricExclusion(metric_id=metric.id, employee_id=int(emp_id))
                db.session.add(exclusion)

        db.session.commit()
        flash("Метрика обновлена.", "success")
        return redirect(url_for('views.admin_metrics'))
    excluded_positions = [excl.position for excl in metric.exclusions if excl.position]
    excluded_employees = [excl.employee for excl in metric.exclusions if excl.employee]

    breadcrumbs = [("Главная", url_for('views.index')), ("Редактировать метрику", "")]
    return render_with_breadcrumbs('edit_metric.html', breadcrumbs,
                                   metric=metric,
                                   categories=categories,
                                   departments=departments,
                                   all_positions=Position.query.all(),
                                   all_employees=Employee.query.filter_by(is_active=True).all(),
                                   excluded_positions=excluded_positions,
                                   excluded_employees=excluded_employees)


@views.route('/admin/metrics/delete/<int:metric_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_metric(metric_id):
    metric = PerformanceMetric.query.get_or_404(metric_id)

    if (current_user.role.name == 'manager' and
        metric.department_id and
        metric.department_id != current_user.department_id):
        abort(403)

    db.session.delete(metric)
    db.session.commit()
    flash("Метрика удалена.", "success")
    return redirect(url_for('views.admin_metrics'))


# --- Управление циклами ---
@views.route('/admin/cycles/add', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_cycle():
    if request.method == 'POST':
        name = request.form.get('name')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        description = request.form.get('description')
        is_active = bool(request.form.get('is_active'))

        if not all([name, start_date, end_date]):
            flash("Заполните название, дату начала и окончания.", "error")
        else:
            try:
                start = datetime.fromisoformat(start_date)
                end = datetime.fromisoformat(end_date)
                if start >= end:
                    flash("Дата начала должна быть раньше даты окончания.", "error")
                else:
                    cycle = EvaluationCycle(
                        name=name,
                        start_date=start,
                        end_date=end,
                        description=description,
                        is_active=is_active
                    )
                    db.session.add(cycle)
                    db.session.commit()
                    flash("Цикл оценки создан.", "success")
                    return redirect(url_for('views.admin_cycles'))
            except ValueError:
                flash("Некорректный формат даты.", "error")

    breadcrumbs = [("Главная", url_for('views.index')), ("Создать цикл", "")]
    return render_with_breadcrumbs('add_cycle.html', breadcrumbs)


@views.route('/admin/cycles/edit/<int:cycle_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_cycle(cycle_id):
    cycle = EvaluationCycle.query.get_or_404(cycle_id)

    if request.method == 'POST':
        new_is_active = bool(request.form.get('is_active'))

        if cycle.is_active and not new_is_active:
            cycle.is_active = False
            db.session.commit()
            flash("Цикл успешно деактивирован.", "success")
            return redirect(url_for('views.admin_cycles'))
        elif cycle.is_active:
            flash("Нельзя изменять параметры активного цикла. Можно только деактивировать.", "error")
            return redirect(url_for('views.admin_cycles'))

        cycle.name = request.form.get('name')
        cycle.start_date = datetime.fromisoformat(request.form.get('start_date'))
        cycle.end_date = datetime.fromisoformat(request.form.get('end_date'))
        cycle.description = request.form.get('description')
        cycle.is_active = new_is_active

        if cycle.start_date >= cycle.end_date:
            flash("Дата начала должна быть раньше даты окончания.", "error")
        else:
            db.session.commit()
            flash("Цикл обновлён.", "success")
            return redirect(url_for('views.admin_cycles'))

    breadcrumbs = [("Главная", url_for('views.index')), (" ", "")]
    return render_with_breadcrumbs('edit_cycle.html', breadcrumbs, cycle=cycle)


@views.route('/admin/cycles/delete/<int:cycle_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_cycle(cycle_id):
    cycle = EvaluationCycle.query.get_or_404(cycle_id)
    if cycle.is_active:
        flash("Нельзя удалить активный цикл.", "error")
    else:
        db.session.delete(cycle)
        db.session.commit()
        flash("Цикл удалён.", "success")
    return redirect(url_for('views.admin_cycles'))

@views.route('/evaluate/<int:emp_id>', methods=['GET', 'POST'])
@login_required
def evaluate(emp_id):
    employee = Employee.query.get_or_404(emp_id)
    cycle = EvaluationCycle.query.filter_by(is_active=True).first()
    
    if not cycle:
        flash("Нет активного цикла оценки.", "error")
        return redirect(url_for('views.dashboard'))

    if current_user.role.name == 'admin':
        pass
    elif current_user.role.name == 'manager':
        if employee.department_id != current_user.department_id:
            flash("Вы можете оценивать только сотрудников своего подразделения.", "error")
            return redirect(url_for('views.dashboard'))
    else:
        if employee.id != current_user.id:
            flash("Вы можете оценить только себя.", "error")
            return redirect(url_for('views.dashboard'))

    emp_position_id = employee.position_id
    excluded_metric_ids = db.session.query(MetricExclusion.metric_id).filter(
        db.or_(
            MetricExclusion.employee_id == emp_id,
            MetricExclusion.position_id == emp_position_id
        )
    ).subquery()

    metrics = PerformanceMetric.query.filter(
        (PerformanceMetric.department_id == employee.department_id) |
        (PerformanceMetric.department_id.is_(None))
    ).filter_by(is_active=True)\
    .filter(~PerformanceMetric.id.in_(excluded_metric_ids))\
    .all()

    if request.method == 'POST':
        for metric in metrics:
            score = request.form.get(f"score_{metric.id}")
            comment = request.form.get(f"comment_{metric.id}")
            if score:
                try:
                    score = float(score)
                    if 0 <= score <= metric.max_score:
                        existing = EmployeeMetric.query.filter_by(
                            employee_id=employee.id,
                            metric_id=metric.id,
                            cycle_id=cycle.id,
                            evaluator_id=current_user.id
                        ).first()
                        if existing:
                            existing.score = score
                            existing.comment = comment
                        else:
                            eval_metric = EmployeeMetric(
                                employee_id=employee.id,
                                metric_id=metric.id,
                                cycle_id=cycle.id,
                                score=score,
                                comment=comment,
                                evaluator_id=current_user.id
                            )
                            db.session.add(eval_metric)
                    else:
                        flash(f"Балл за '{metric.name}' вне диапазона.", "error")
                except ValueError:
                    flash(f"Некорректное значение для '{metric.name}'.", "error")
        db.session.commit()
        flash(f"Оценка сотрудника {employee.full_name} сохранена.", "success")
        return redirect(url_for('views.dashboard'))

    breadcrumbs = [("Главная", url_for('views.index')), ("Оценка", "")]
    return render_with_breadcrumbs('evaluate.html', breadcrumbs,
                                   employee=employee,
                                   metrics=metrics,
                                   cycle=cycle)

@views.route('/export-import')
@login_required
@admin_or_manager_required
def export_import():
    if current_user.role.name not in ['manager', 'admin']:
        abort(403)

    breadcrumbs = [
        ("Главная", url_for('views.index')),
        ("Импорт/Экспорт", "")
    ]
    return render_with_breadcrumbs('export_import.html', breadcrumbs)

@views.route('/import-data', methods=['POST'])
@login_required
@admin_or_manager_required
def import_data():
    if 'file' not in request.files:
        flash("Файл не выбран.", "error")
        return redirect(url_for('views.export_import'))

    file = request.files['file']
    if file.filename == '':
        flash("Файл не выбран.", "error")
        return redirect(url_for('views.export_import'))

    # Проверка расширения
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        flash("Поддерживаются только .xlsx и .xls файлы.", "error")
        return redirect(url_for('views.export_import'))

    active_cycle = EvaluationCycle.query.filter_by(is_active=True).first()
    if not active_cycle:
        flash("Нет активного цикла оценки.", "error")
        return redirect(url_for('views.export_import'))

    try:
        # Сохраним файл временно
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.tmp') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
    
        imported = 0
    
        # --- НОВЫЙ БЛОК: Определение типа файла и чтение ---
        if file.filename.lower().endswith('.csv'):
            import csv
            with open(tmp_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)
                if len(rows) < 2:
                    flash("Файл пуст или содержит только заголовок.", "error")
                    return redirect(url_for('views.export_import'))
                # Пропускаем первую строку (заголовок)
                for row in rows[1:]:
                    if len(row) < 3:
                        continue
                    emp_email, metric_name, score = row[0], row[1], row[2]
                    comment = row[3] if len(row) > 3 else ''
                    # ... остальная логика обработки строки (как в Excel) ...
                    employee = Employee.query.filter_by(email=emp_email).first()
                    metric = PerformanceMetric.query.filter_by(name=metric_name).first()
                    if not employee:
                        flash(f"Не найден сотрудник с email: {emp_email}", "warning")
                        continue
                    if not metric:
                        flash(f"Не найдена метрика: {metric_name}", "warning")
                        continue
                    try:
                        score = float(score)
                    except (TypeError, ValueError):
                        flash(f"Некорректный балл для {emp_email}: {score}", "error")
                        continue
                    existing = EmployeeMetric.query.filter_by(
                        employee_id=employee.id,
                        metric_id=metric.id,
                        cycle_id=active_cycle.id
                    ).first()
                    if existing:
                        existing.score = score
                        existing.comment = str(comment)
                    else:
                        eval_metric = EmployeeMetric(
                            employee_id=employee.id,
                            metric_id=metric.id,
                            cycle_id=active_cycle.id,
                            score=score,
                            comment=str(comment),
                            evaluator_id=current_user.id
                        )
                        db.session.add(eval_metric)
                    imported += 1
    
        else:  # Это Excel
            from openpyxl import load_workbook
            workbook = load_workbook(tmp_path)
            sheet = workbook.active
            if sheet.max_row < 2:
                flash("Файл пуст или содержит только заголовок.", "error")
                return redirect(url_for('views.export_import'))
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                emp_email, metric_name, score, comment = row[0], row[1], row[2], row[3] or ''
                # ... та же логика обработки строки, что и выше ...
                employee = Employee.query.filter_by(email=emp_email).first()
                metric = PerformanceMetric.query.filter_by(name=metric_name).first()
                if not employee:
                    flash(f"Не найден сотрудник с email: {emp_email}", "warning")
                    continue
                if not metric:
                    flash(f"Не найдена метрика: {metric_name}", "warning")
                    continue
                try:
                    score = float(score)
                except (TypeError, ValueError):
                    flash(f"Некорректный балл для {emp_email}: {score}", "error")
                    continue
                existing = EmployeeMetric.query.filter_by(
                    employee_id=employee.id,
                    metric_id=metric.id,
                    cycle_id=active_cycle.id
                ).first()
                if existing:
                    existing.score = score
                    existing.comment = str(comment)
                else:
                    eval_metric = EmployeeMetric(
                        employee_id=employee.id,
                        metric_id=metric.id,
                        cycle_id=active_cycle.id,
                        score=score,
                        comment=str(comment),
                        evaluator_id=current_user.id
                    )
                    db.session.add(eval_metric)
                imported += 1
    
        db.session.commit()
        flash(f"Успешно импортировано {imported} оценок.", "success")
    
    except Exception as e:
        db.session.rollback()
        print("Ошибка при импорте:", str(e))
        flash(f"Ошибка при обработке файла: {str(e)}", "error")

    except Exception as e:
        db.session.rollback()
        print("Ошибка при импорте:", str(e))
        flash(f"Ошибка при обработке файла: {str(e)}", "error")

    return redirect(url_for('views.export_import'))

@views.route('/export-pdf', methods=['POST'])
@login_required
@admin_or_manager_required
def export_pdf():
    report_type = request.form.get('report_type')
    department_id = request.form.get('department_id')

    active_cycle = EvaluationCycle.query.filter_by(is_active=True).first()
    if not active_cycle:
        flash("Нет активного цикла оценки.", "error")
        return redirect(request.url)

    from sqlalchemy import func

    # Получаем средний балл
    avg_score = db.session.query(func.avg(EmployeeMetric.score)).filter(
        EmployeeMetric.cycle_id == active_cycle.id
    ).scalar()
    avg_score = avg_score or 0

    query = db.session.query(
        Employee.full_name,
        Department.name.label('dept_name'),
        func.avg(EmployeeMetric.score).label('avg_score')
    ).join(EmployeeMetric, Employee.id == EmployeeMetric.employee_id) \
     .join(Department, Employee.department_id == Department.id) \
     .filter(EmployeeMetric.cycle_id == active_cycle.id) \
     .group_by(Employee.id, Department.name)
    
    if report_type == 'above_avg':
        query = query.having(func.avg(EmployeeMetric.score) > avg_score)
        title = "Сотрудники с оценкой выше средней"
    elif report_type == 'below_avg':
        query = query.having(func.avg(EmployeeMetric.score) < avg_score)
        title = "Сотрудники с оценкой ниже средней"
    elif report_type == 'by_department' and department_id:
        query = query.filter(Employee.department_id == int(department_id))
        dept_name = Department.query.get(department_id).name
        title = f"Сотрудники подразделения: {dept_name}"
    else:
        title = "Все сотрудники"

    employees = query.order_by(func.avg(EmployeeMetric.score).desc()).all()

    pdf = FPDF()
    pdf.add_font("DejaVu", style="", fname=FONT_PATH)
    pdf.add_font("DejaVu", style="B", fname=FONT_PATH_B)
    pdf.add_font("DejaVu", style="I", fname=FONT_PATH_I)
    pdf.set_font("DejaVu", size=12)
    pdf.add_page()

    # Заголовок
    pdf.cell(0, 10, 'Отчёт по эффективности', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.cell(0, 10, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(5)

    # Информация
    pdf.cell(0, 8, f"Цикл: {active_cycle.name}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 8, f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(10)

    # Шапка таблицы
    pdf.set_font("DejaVu", 'B', 12)
    col_widths = [80, 60, 30]
    pdf.set_fill_color(200, 220, 255)
    pdf.cell(col_widths[0], 10, 'ФИО', border=1, fill=True)
    pdf.cell(col_widths[1], 10, 'Подразделение', border=1, fill=True)
    pdf.cell(col_widths[2], 10, 'Ср. балл', border=1, fill=True)
    pdf.ln(10)

    # Данные
    pdf.set_font("DejaVu", size=11)
    for emp in employees:
        pdf.cell(col_widths[0], 8, emp.full_name, border=1)
        pdf.cell(col_widths[1], 8, emp.dept_name, border=1)
        pdf.cell(col_widths[2], 8, f"{emp.avg_score:.2f}", border=1)
        pdf.ln(8)

    pdf.ln(10)
    pdf.set_font("DejaVu", 'I', 10)
    pdf.cell(0, 8, f"Всего: {len(employees)} сотруд.", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf_output = pdf.output()

    return send_file(
        BytesIO(pdf_output),
        as_attachment=True,
        download_name=f'report_{report_type}_{datetime.now().strftime("%d%m%Y")}.pdf',
        mimetype='application/pdf'
    )

@views.route('/admin/messages')
@login_required
@admin_or_manager_required
def admin_messages():
    if current_user.role.name != 'admin':
        abort(403)
    messages = ContactMessage.query.order_by(ContactMessage.created_at.desc()).all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Сообщения обратной связи", "")]
    return render_with_breadcrumbs('admin_messages.html', breadcrumbs, messages=messages)


@views.route('/admin/messages/toggle-read/<int:message_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def toggle_message_read(message_id):
    if current_user.role.name != 'admin':
        abort(403)
    message = ContactMessage.query.get_or_404(message_id)
    message.read = not message.read
    db.session.commit()
    return redirect(url_for('views.admin_messages'))


@views.route('/admin/messages/delete/<int:message_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_message(message_id):
    if current_user.role.name != 'admin':
        abort(403)
    message = ContactMessage.query.get_or_404(message_id)
    db.session.delete(message)
    db.session.commit()
    flash("Сообщение удалено.", "success")
    return redirect(url_for('views.admin_messages'))
    
    
# --- Управление должностями ---
@views.route('/admin/positions')
@login_required
@admin_or_manager_required
def admin_positions():
    if current_user.role.name == 'manager':
        positions_list = Position.query.filter_by(department_id=current_user.department_id).all()
    else:
        positions_list = Position.query.all()
    breadcrumbs = [("Главная", url_for('views.index')), ("Управление должностями", "")]
    return render_with_breadcrumbs('admin_positions.html', breadcrumbs, positions=positions_list)


@views.route('/admin/positions/add', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def add_position():
    departments = [current_user.department] if current_user.role.name == 'manager' else Department.query.all()

    if request.method == 'POST':
        title = request.form.get('title')
        department_id = request.form.get('department_id')
        description = request.form.get('description')

        if not all([title, department_id]):
            flash("Заполните название и подразделение.", "error")
        else:
            if current_user.role.name == 'manager' and int(department_id) != current_user.department_id:
                flash("Вы можете добавлять должности только в своё подразделение.", "error")
                return redirect(url_for('views.add_position'))

            position = Position(
                title=title,
                department_id=int(department_id),
                description=description
            )
            db.session.add(position)
            db.session.commit()
            flash(f"Должность '{title}' добавлена.", "success")
            return redirect(url_for('views.admin_positions'))

    breadcrumbs = [("Главная", url_for('views.index')), ("Добавить должность", "")]
    return render_with_breadcrumbs('add_position.html', breadcrumbs, departments=departments)


@views.route('/admin/positions/edit/<int:pos_id>', methods=['GET', 'POST'])
@login_required
@admin_or_manager_required
def edit_position(pos_id):
    position = Position.query.get_or_404(pos_id)

    if current_user.role.name == 'manager' and position.department_id != current_user.department_id:
        abort(403)

    departments = [current_user.department] if current_user.role.name == 'manager' else Department.query.all()

    if request.method == 'POST':
        title = request.form.get('title')
        department_id = request.form.get('department_id')
        description = request.form.get('description')

        if not all([title, department_id]):
            flash("Заполните название и подразделение.", "error")
        else:
            if current_user.role.name == 'manager' and int(department_id) != current_user.department_id:
                flash("Вы можете редактировать должности только в своём подразделении.", "error")
                return redirect(url_for('views.edit_position', pos_id=pos_id))

            position.title = title
            position.department_id = int(department_id)
            position.description = description
            db.session.commit()
            flash("Должность обновлена.", "success")
            return redirect(url_for('views.admin_positions'))

    breadcrumbs = [("Главная", url_for('views.index')), ("Редактировать должность", "")]
    return render_with_breadcrumbs('edit_position.html', breadcrumbs, position=position, departments=departments)


@views.route('/admin/positions/delete/<int:pos_id>', methods=['POST'])
@login_required
@admin_or_manager_required
def delete_position(pos_id):
    position = Position.query.get_or_404(pos_id)

    if current_user.role.name == 'manager' and position.department_id != current_user.department_id:
        abort(403)

    emp_count = Employee.query.filter_by(position_id=pos_id).count()
    if emp_count > 0:
        flash(f"Нельзя удалить должность '{position.title}', так как на ней числится {emp_count} сотрудников.", "error")
    else:
        db.session.delete(position)
        db.session.commit()
        flash("Должность удалена.", "success")

    return redirect(url_for('views.admin_positions'))